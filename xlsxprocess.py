import openpyxl as op
import argparse
import re
import pdb

######Convert cursor string to duple####################
def cursortoints(string_cursor):
	#Todo: columns with 2 letters
	column=re.search('[A-Z]',string_cursor)
	row = re.search('[0-9]+',string_cursor)
	
	#columns start at 1
	column=ord(column.group(0))-64
	row = int(row.group(0))
	return column, row
	
####Convert cursor duple to string######################
def cursortostring((column, row)):
	cursor_string =""
	cursor_string= chr(column + 64) + str(row)
	return cursor_string


	
#####return range when given integer coordinates of top left and bot right
def returnrange(ws,topleft,bottomright):
	return ws.range(cursortostring(topleft)+':'+
						cursortostring(bottomright))

#################Remove headers#########################
#Puts the cursor at the beginning of the data
def removeheaders(ws,cursor):
	
	#Search for "Contents in Column A"
	for x in ws.rows:
		if x[0].value == "Contents":
			cursor= cursortoints(x[0].address)
			#increment cursor row to get past header once keyword found
			cursor=(cursor[0], cursor[1]+1)
	#If not found return existing cursor
	return cursor
			
#########search on the cursor row for the string########
def lookright(ws,cursor,string):
	#Look in columns A->Z, can't do anything past that yet
	cells=returnrange(ws,cursor,(cursor[0]+25,cursor[1]))[0]

	for cell in cells:
		if cell.value !=None or "":
			if cell.value.find(string)!=-1:
				return cell.value
	
	return None

#########Look for coach string #####
def lookforcoach(ws,cursor):
	coachstring=lookright(ws,cursor,'Coach')
	return coachstring[7:]

#########Look for manager string #####
def lookformanager(ws,cursor):
	managerstring=lookright(ws,cursor,'Manager')
	
	return managerstring[9:]
	
	
#########Look for State Cup Result string #####
def lookforstatecup(ws,cursor):
	statecupstring=lookright(ws,cursor,'State Cup')
	
	return statecupstring

#########return cell contents when given integer cursor####
def cellcontents(ws,cursor):
	return ws.cell(cursortostring((cursor[0],cursor[1]))).value


##########look for next player and return address of player number cell 
def lookfornextplayer(ws,cursor):
	for row in ws.range(cursortostring((cursor[0],cursor[1]+1))+':'+cursortostring(
						(cursor[0],cursor[1]+10))):
			#print "Looking for next player"
			#print row
			#print row[0].value
			if row[0].value != None:
				m=re.search('#',row[0].value)
				if m != None:
					if m.group(0) == "#":
						return cursortoints(row[0].address)
	return cursor

def validateemail(string):

	if string==None:
		return None
	m = re.search('@',string)
	if m != None:
		if m.group(0)=='@':
			return string
	
	return None

def getplayerdata(ws,cursor):
	player=dict()
	player['number']=cellcontents(ws,cursor)
	player['name']=cellcontents(ws,(cursor[0],cursor[1]+1))
	player['DOB']=lookright(ws,(cursor[0],cursor[1]),'DOB')
	if player['DOB']==None:
		player['DOB']=lookright(ws,(cursor[0],cursor[1]+1),'DOB')
	player['position']=lookright(ws,(cursor[0],cursor[1]),',')
	
	player['email']=validateemail(cellcontents(ws,(cursor[0],cursor[1]+2)))
	player['phone']=lookright(ws,(cursor[0],cursor[1]+1),'Phone')
	if player['phone'] == None:
		player['phone']=lookright(ws,(cursor[0],cursor[1]+2),'Phone')
	if player['phone'] == None:
		player['phone']=lookright(ws,(cursor[0],cursor[1]+3),'Phone')	
	player['grad']=lookright(ws,(cursor[0],cursor[1]+1),'Grad Year')
		
	return player

def lastplayeronteam(ws,cursor):
	for row in ws.range(cursortostring((cursor[0],cursor[1]+1))+':'+cursortostring(
					(cursor[0],cursor[1]+10))):
		#print "Looking for last player on the team"
		#print row
		#print row[0].value
		if row[0].value != None:
			m=re.search('Event Schedule',row[0].value)
			if m != None:
				if m.group(0) == "Event Schedule":
					return True
			m=re.search('State Cup',row[0].value)
			if m != None:
				if m.group(0) == "State Cup":
					return True
			m=re.search('#',row[0].value)
			if m != None:
				if m.group(0) == "#":
					return False
	return False

def teamrosternotavailable(ws,cursor):
	for row in ws.range(cursortostring((cursor[0],cursor[1]+1))+':'+cursortostring(
					(cursor[0],cursor[1]+1))):
		#print "Looking if roster is available"
		#print row
		#print row[0].value
		if row[0].value != None:
			m=re.search('Team Roster Not Available',row[0].value)
			if m != None:
				if m.group(0) == "Team Roster Not Available":
					return True
	return False

def findnextteam(ws,cursor):
	for row in ws.range(cursortostring((cursor[0],cursor[1]+1))+':'+cursortostring(
					(cursor[0],cursor[1]+10))):
		#print "Looking if next team is available"
		#print row
		#print row[0].value
		if row[0].value != None:
			m=re.search('Event Schedule',row[0].value)
			if m != None:
				if m.group(0) == "Event Schedule":
					return None
			m=re.search('State Cup',row[0].value)
			if m != None:
				if m.group(0) == "State Cup":
					return cursortoints(row[0].address)
	return None

def setupoutputws(wsout,cursorout):
	#print "Setting up output WS"
	c=wsout.cell(cursortostring(cursorout))
	c.value="Player"
	c=wsout.cell(cursortostring((cursorout[0]+1,cursorout[1])))
	c.value="Number"
	c=wsout.cell(cursortostring((cursorout[0]+2,cursorout[1])))
	c.value="Email"
	c=wsout.cell(cursortostring((cursorout[0]+3,cursorout[1])))
	c.value="DOB"
	c=wsout.cell(cursortostring((cursorout[0]+4,cursorout[1])))
	c.value="Phone"
	
	return (cursorout[0],cursorout[1]+1)

#################Main#######################

#Parsing of arguments
parser = argparse.ArgumentParser(description='Process Joe\'s xlsx')
parser.add_argument('-f','--xlsxfile', help='File to be processed', required=True)
args = vars(parser.parse_args())

#Loading the workbook then grab default worksheet
wb = op.load_workbook(filename=args['xlsxfile'],use_iterators=False)
ws = wb.get_active_sheet()

wbout=op.Workbook()
wsout=wbout.get_active_sheet()


#The cursor is to keep track of where we are in the file
cursor=tuple()
cursorout=(1,1)
cursorout=setupoutputws(wsout,cursorout)
#print cursorout
cursor=removeheaders(ws,cursor)
teams=[]
while True:
	coach=""
	manager=""
	statecup=""
	players=[]

	coach=lookforcoach(ws,cursor)
	#manager=lookformanager(ws,cursor)
	statecup=lookforstatecup(ws,cursor)

	while True:
		if teamrosternotavailable(ws,cursor):
			break
		cursor=lookfornextplayer(ws,cursor)

		players.append(getplayerdata(ws,cursor))
		if lastplayeronteam(ws,cursor):
			break

	#print "Coach: %s"%coach
	#print "Manager: %s"% manager
	#print "State Cup result: %s"% statecup
	#print players
	#print cursor
	#print cursortostring(cursor)
	cursor = findnextteam(ws,cursor)
	if cursor == None:
		break
	#print cursor
	#print cursortostring(cursor)

	for player in players:

		c = wsout.cell(cursortostring(cursorout))
		c.value=player['name']
		c = wsout.cell(cursortostring(cursorout))
		c=wsout.cell(cursortostring((cursorout[0]+1,cursorout[1])))
		c.value=player['number']
		c=wsout.cell(cursortostring((cursorout[0]+2,cursorout[1])))
		c.value=player['email']
		c=wsout.cell(cursortostring((cursorout[0]+3,cursorout[1])))
		c.value=player['DOB']
		c=wsout.cell(cursortostring((cursorout[0]+4,cursorout[1])))
		c.value=player['phone']
		cursorout=(cursorout[0],cursorout[1]+1)
		#print cursorout


wbout.save('out.xlsx')
